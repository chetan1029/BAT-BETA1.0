import json
import tempfile
import csv
from io import StringIO
from decimal import Decimal


from django.http import HttpResponse

from rest_framework import status
from measurement.measures import Weight


import openpyxl
from bat.product.models import Product
from bat.setting.utils import get_status
from bat.product.constants import PRODUCT_STATUS, PRODUCT_PARENT_STATUS


class ProductCSVErrorBuilder(object):

    @classmethod
    def write(cls, invalid_records, data_file):

        # header mapping
        headers = [*invalid_records[0]]
        try:
            # create csv file
            csvfile = StringIO()
            dict_writer = csv.DictWriter(csvfile, headers, extrasaction='ignore')
            dict_writer.writeheader()
            dict_writer.writerows(invalid_records)
            csvfile.seek(0)
            return csvfile
        except Exception:
            return None


class ProductCSVParser(object):

    @classmethod
    def parse(cls, csv_file):
        data = []
        # read file
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file, delimiter=',')

        line_number = 2
        for row in reader:
            row.pop("id")
            values = {"line_number": line_number}
            line_number += 1

            for key, value in row.items():
                if key in ["length", "width", "depth"] and value == "":
                    value = None
                elif key == "status":
                    values[key] = get_status(PRODUCT_PARENT_STATUS, PRODUCT_STATUS.get(
                        value.lower())) if value != "" else None
                elif key == "weight":
                    value = value.replace(" g", "") if value != "" else None
                    value = Weight({"g": Decimal(value)})
                    values[key] = value
                elif key == "tags":
                    values[key] = value.split(",") if value != "" else None
                elif key == "errors":
                    pass
                else:
                    values[key] = value
            data.append(values)
        header = reader.fieldnames
        header.remove("id")
        return data, header


class ProductExcelErrorBuilder(object):

    @classmethod
    def write(cls, invalid_records, data_file):
        # Temporary files
        tmp_dir = tempfile.TemporaryDirectory()
        tmp_xlsx_file_path = tmp_dir.name + "/" + data_file.name

        # create Workbook
        wb = openpyxl.load_workbook(data_file)
        sheet = wb.active

        # header mapping
        lat_col_index = len(sheet[1])+1
        header_cell = sheet.cell(row=1, column=lat_col_index)
        header_cell.value = "errors"

        try:
            # write data
            for row in invalid_records:
                row_number = row.get("line_number")
                error_cell = sheet.cell(row=row_number, column=lat_col_index)
                error_cell.value = row.get("errors")
            wb.save(tmp_xlsx_file_path)
            wb.close()
            f = open(tmp_xlsx_file_path, "rb")
            return f
        except Exception:
            return None


class ProductExcelParser(object):

    @classmethod
    def parse(cls, excel_file):
        data = []
        # read file
        ws = openpyxl.load_workbook(excel_file)
        sheet = ws.active

        # get column's list
        header = [cell.value for cell in sheet[1]]
        header.remove("id")

        for row in sheet.iter_rows(min_row=2, min_col=2):
            values = {"line_number": row[0].row}
            # process each value
            for key, cell in zip(header, row):
                value = cell.value
                if key == "manufacturer_part_number":
                    values[key] = value if value else ""
                elif key == "status":
                    values[key] = get_status(PRODUCT_PARENT_STATUS, PRODUCT_STATUS.get(
                        value.lower())) if value else None
                elif key == "weight":
                    value = value.replace(" g", "") if value else None
                    value = Weight({"g": Decimal(value)})
                    values[key] = value
                elif key == "tags":
                    values[key] = value.split(",") if value else None
                elif key == "hscode":
                    values[key] = value if value else ""
                elif key in ["bullet_points", "description"]:
                    values[key] = value if value else ""
                elif key == "errors":
                    pass
                else:
                    values[key] = value
            data.append(values)
        return data, header
